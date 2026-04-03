#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import contextlib
import http.server
import json
import os
import re
import shutil
import socketserver
import tempfile
import threading
import time
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError


# -----------------------------
# Config
# -----------------------------

@dataclass(frozen=True)
class MarketConfig:
    market_id: str
    market_label: str
    group_label: str
    outcome_label: str
    prob_min: int
    odd_min: float = 1.01
    odd_max: float = 10.0
    notes: str = ""


MARKETS: List[MarketConfig] = [
    MarketConfig("over_25", "Over 2.5", "Goal", "Over 2.5", 80, 1.01, 10.0, "80%+ qualsiasi quota"),
    MarketConfig("over_15", "Over 1.5", "Goal", "Over 1.5", 80, 1.30, 10.0, "80%+ quota >= 1.30"),
    MarketConfig("over_35", "Over 3.5", "Goal", "Over 3.5", 75, 1.01, 10.0, "75%+ qualsiasi quota"),
    MarketConfig("over_15_ht", "Over 1.5 HT", "1T / 2T", "Over 1.5 HT", 70, 1.01, 10.0, "70%+ qualsiasi quota"),
    MarketConfig("dc_1x", "1X", "1X2 / DC", "1X", 70, 1.30, 10.0, "70%+ quota >= 1.30"),
    MarketConfig("dc_x2", "X2", "1X2 / DC", "X2", 70, 1.30, 10.0, "70%+ quota >= 1.30"),
    MarketConfig("btts_yes", "BTTS YES", "Entrambe segnano", "BTTS YES", 80, 1.01, 10.0, "80%+ qualsiasi quota"),
    MarketConfig("btts_no", "BTTS NO", "Entrambe segnano", "BTTS NO", 80, 1.01, 10.0, "80%+ qualsiasi quota"),
    MarketConfig("corner_over_85", "Corner Over 8.5", "Corner", "Over 8.5", 80, 1.01, 10.0, "80%+ qualsiasi quota"),
    MarketConfig("corner_under_105", "Corner Under 10.5", "Corner", "Under 10.5", 80, 1.01, 10.0, "80%+ qualsiasi quota"),
    # Il frontend attuale espone linee yellows over 2.5/3.5/4.5: per 'over 4 cards' usiamo Over 4.5.
    MarketConfig("cards_over_4", "Over 4 Cards", "Cartellini", "Over 4.5", 80, 1.01, 10.0, "80%+ qualsiasi quota (mappato a Over 4.5)"),
    # Il frontend espone 'Over 2.5 + BTTS'.
    MarketConfig("combo_o25_btts_yes", "Combo Over 2.5 + BTTS Yes", "Combo", "Over 2.5 + BTTS", 80, 1.01, 10.0, "80%+ qualsiasi quota"),
]

PAGE_CANDIDATES = ["/match-center.html", "/index.html", "/"]
SHEET_ORDER = [
    "summary",
    "final_selected",
    "all_raw",
    "over_25",
    "over_15",
    "over_35",
    "over_15_ht",
    "dc_1x",
    "dc_x2",
    "btts_yes",
    "btts_no",
    "corner_over_85",
    "corner_under_105",
    "cards_over_4",
    "combo_o25_btts_yes",
    "debug",
]


# -----------------------------
# Small helpers
# -----------------------------


def safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(",", ".")
        return float(text)
    except Exception:
        return default



def jsonable(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, dict):
        return {str(k): jsonable(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [jsonable(v) for v in value]
    return str(value)



def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)



def ddmm_from_iso(iso_date: str) -> str:
    try:
        parts = iso_date.split("-")
        return f"{parts[2]}/{parts[1]}"
    except Exception:
        return iso_date



def current_ts() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")



def unique_match_key(row: Dict[str, Any]) -> str:
    home = str(row.get("home") or "").strip().lower()
    away = str(row.get("away") or "").strip().lower()
    date = str(row.get("date") or "").strip()
    time_s = str(row.get("time") or "").strip()
    league = str(row.get("league") or "").strip().lower()
    return "|".join([date, time_s, league, home, away])


class QuietHTTPRequestHandler(http.server.SimpleHTTPRequestHandler):
    def log_message(self, format: str, *args: Any) -> None:
        return


class ReusableTCPServer(socketserver.TCPServer):
    allow_reuse_address = True


@contextlib.contextmanager
def serve_directory(root: Path):
    handler = QuietHTTPRequestHandler
    cwd = os.getcwd()
    os.chdir(str(root))
    try:
        with ReusableTCPServer(("127.0.0.1", 0), handler) as httpd:
            port = httpd.server_address[1]
            thread = threading.Thread(target=httpd.serve_forever, daemon=True)
            thread.start()
            try:
                yield f"http://127.0.0.1:{port}"
            finally:
                httpd.shutdown()
                thread.join(timeout=5)
    finally:
        os.chdir(cwd)


# -----------------------------
# Workspace preparation
# -----------------------------


def prepare_workspace(repo_root: Path, target_date: str) -> Path:
    temp_root = Path(tempfile.mkdtemp(prefix="mc-ui-dry-run-"))
    work_root = temp_root / "site"
    shutil.copytree(repo_root, work_root, dirs_exist_ok=True)

    cache_dir = work_root / "assets" / "data" / "match-predictor"
    date_json = cache_dir / f"{target_date}.json"
    latest_json = cache_dir / "latest.json"
    if date_json.exists():
        shutil.copy2(date_json, latest_json)

    date_standings = cache_dir / f"{target_date}.standings.json"
    latest_standings = cache_dir / "latest.standings.json"
    if date_standings.exists():
        shutil.copy2(date_standings, latest_standings)

    manifest_path = cache_dir / "manifest.json"
    if manifest_path.exists():
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            if isinstance(manifest, dict):
                dates = manifest.get("dates")
                if isinstance(dates, list):
                    for entry in dates:
                        if isinstance(entry, dict) and entry.get("date") == target_date:
                            manifest["latest"] = target_date
                            break
                manifest_path.write_text(json.dumps(manifest, ensure_ascii=False), encoding="utf-8")
        except Exception:
            pass

    return work_root


# -----------------------------
# Playwright UI helpers
# -----------------------------


def open_match_center(page, base_url: str) -> str:
    for candidate in PAGE_CANDIDATES:
        url = f"{base_url}{candidate}"
        try:
            page.goto(url, wait_until="networkidle", timeout=90000)
            page.wait_for_timeout(1200)
            body_text = page.locator("body").inner_text(timeout=15000)
            if "Probabilità" in body_text and "Pronostici" in body_text:
                return url
        except PlaywrightTimeoutError:
            continue
        except Exception:
            continue
    raise RuntimeError("Match Center non trovato nelle pagine candidate")



def wait_for_ui_ready(page) -> None:
    page.wait_for_timeout(1500)
    page.wait_for_function(
        """
        () => {
          const text = document.body.innerText || "";
          return text.includes("Probabilità") && text.includes("Pronostici");
        }
        """,
        timeout=30000,
    )
    page.wait_for_timeout(1200)



def click_first_matching_button(page, labels: Iterable[str], timeout: int = 8000) -> bool:
    for label in labels:
        try:
            locator = page.get_by_role("button", name=re.compile(re.escape(label), re.IGNORECASE)).first
            locator.wait_for(timeout=timeout)
            locator.click(timeout=timeout)
            page.wait_for_timeout(500)
            return True
        except Exception:
            continue
    return False



def set_select_by_context(page, context_label: str, desired: str) -> bool:
    script = """
    ({ contextLabel, desired }) => {
      const visible = (el) => {
        if (!el) return false;
        const style = window.getComputedStyle(el);
        const rect = el.getBoundingClientRect();
        return style && style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
      };
      const norm = (s) => String(s || '').toLowerCase().replace(/\s+/g, ' ').trim();
      const want = norm(desired);
      const label = norm(contextLabel);
      const selects = [...document.querySelectorAll('select')].filter(visible);
      for (const sel of selects) {
        let node = sel;
        let matched = false;
        for (let i = 0; i < 4 && node; i += 1, node = node.parentElement) {
          const text = norm(node.innerText || node.textContent || '');
          if (text.includes(label)) {
            matched = true;
            break;
          }
        }
        if (!matched) continue;
        const options = [...sel.options];
        const candidate = options.find(opt => norm(opt.textContent).includes(want) || norm(opt.value).includes(want));
        if (!candidate) continue;
        sel.value = candidate.value;
        sel.dispatchEvent(new Event('input', { bubbles: true }));
        sel.dispatchEvent(new Event('change', { bubbles: true }));
        return true;
      }
      return false;
    }
    """
    try:
        ok = page.evaluate(script, {"contextLabel": context_label, "desired": desired})
        page.wait_for_timeout(600)
        return bool(ok)
    except Exception:
        return False



def set_full_day_filters(page, market: MarketConfig) -> None:
    # quick all day if present
    click_first_matching_button(page, ["Tutto il giorno", "All day"], timeout=2000)

    set_select_by_context(page, "PROB. MIN", f"{market.prob_min}%")
    set_select_by_context(page, "PROB. MAX", "99%")
    set_select_by_context(page, "QUOTA MIN", f"{market.odd_min:.2f}".rstrip("0").rstrip("."))
    set_select_by_context(page, "QUOTA MAX", f"{market.odd_max:.2f}".rstrip("0").rstrip("."))

    # fallback time range if needed
    set_select_by_context(page, "ORARIO", "00:00") or set_select_by_context(page, "ORARIO", "00:30")
    set_select_by_context(page, "QUOTA MAX", "10")
    page.wait_for_timeout(1200)



def apply_market_filter(page, market: MarketConfig) -> None:
    # Most reliable reset is page reload.
    page.reload(wait_until="networkidle", timeout=90000)
    wait_for_ui_ready(page)

    if not click_first_matching_button(page, [market.group_label], timeout=8000):
        raise RuntimeError(f"Bottone gruppo non trovato: {market.group_label}")

    # Often the group click reveals a 'Tutti gli esiti' chip. Not mandatory.
    click_first_matching_button(page, ["Tutti gli esiti", "Tutte le linee", "All outcomes", "All lines"], timeout=1500)

    if not click_first_matching_button(page, [market.outcome_label], timeout=8000):
        raise RuntimeError(f"Bottone esito non trovato: {market.outcome_label}")

    # Probability preset, when present.
    click_first_matching_button(page, [f"{market.prob_min}%+"], timeout=1500)
    set_full_day_filters(page, market)
    page.wait_for_timeout(1800)


EXTRACT_JS = r'''
(marketMeta) => {
  const visible = (el) => {
    if (!el) return false;
    const style = window.getComputedStyle(el);
    const rect = el.getBoundingClientRect();
    return style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 20 && rect.height > 20;
  };
  const norm = (s) => String(s || '').replace(/\u00a0/g, ' ').replace(/\s+/g, ' ').trim();
  const all = [...document.querySelectorAll('div,article,section,li,a')].filter(visible);
  const rawMatches = all.filter(el => {
    const txt = norm(el.innerText || el.textContent || '');
    if (!txt) return false;
    if (!txt.includes(' vs ')) return false;
    if (!/%/.test(txt)) return false;
    if (!/\d{2}:\d{2}/.test(txt)) return false;
    return true;
  });

  const minimal = rawMatches.filter(el => {
    const children = [...el.querySelectorAll('div,article,section,li,a')].filter(visible);
    return !children.some(child => child !== el && rawMatches.includes(child));
  });

  const parse = (text) => {
    const clean = norm(text);
    const lines = clean.split(/\n+/).map(norm).filter(Boolean);
    const timeMatch = clean.match(/\b\d{2}:\d{2}\b/);
    const dateMatch = clean.match(/\b\d{2}\/\d{2}\b/);
    const teamLine = lines.find(line => line.includes(' vs ')) || '';
    const [home, away] = teamLine.split(' vs ').map(norm);
    const leagueLine = lines.find(line => line.includes('|') && !line.includes('%') && !line.includes('vs')) || '';
    const leagueParts = leagueLine.split('|').map(norm).filter(Boolean);
    const percents = [...clean.matchAll(/(\d{1,2}(?:\.\d+)?)%/g)].map(m => Number(m[1])).filter(Number.isFinite);
    const odds = [...clean.matchAll(/\b(1\.(?:\d{1,2})|[2-9](?:\.\d{1,2})?|10(?:\.0+)?)\b/g)].map(m => Number(m[1])).filter(Number.isFinite);
    const probability = percents.length ? Math.max(...percents) : null;
    let odd = null;
    if (odds.length) {
      odd = odds.sort((a, b) => b - a)[0];
    }
    return {
      date: dateMatch ? dateMatch[0] : '',
      time: timeMatch ? timeMatch[0] : '',
      league: leagueParts[0] || '',
      country: leagueParts[1] || '',
      home: home || '',
      away: away || '',
      probability,
      odd,
      raw_text: clean,
    };
  };

  const rows = [];
  const seen = new Set();
  for (const el of minimal) {
    const parsed = parse(el.innerText || el.textContent || '');
    if (!parsed.home || !parsed.away || !parsed.time || parsed.probability == null) continue;
    const key = [parsed.date, parsed.time, parsed.league, parsed.home, parsed.away].join('|').toLowerCase();
    if (!key || seen.has(key)) continue;
    seen.add(key);
    rows.push({
      ...parsed,
      market_id: marketMeta.market_id,
      market_label: marketMeta.market_label,
      source: 'ui_rendered',
      notes: marketMeta.notes || '',
    });
  }

  const countText = norm(document.body.innerText || '');
  const pronosticiMatch = countText.match(/Pronostici\((\d+)\)/i);

  return {
    rows,
    visible_count: rows.length,
    heading_count: pronosticiMatch ? Number(pronosticiMatch[1]) : null,
  };
}
'''


def extract_market_rows(page, market: MarketConfig) -> Dict[str, Any]:
    result = page.evaluate(EXTRACT_JS, asdict(market))
    return result if isinstance(result, dict) else {"rows": [], "visible_count": 0, "heading_count": None}


# -----------------------------
# Output
# -----------------------------


def autofit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        column = col[0].column
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(column)].width = min(max_len + 2, 40)



def append_sheet(ws, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        ws.append(["empty"])
        return
    headers: List[str] = []
    seen = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                headers.append(key)
    ws.append(headers)
    for row in rows:
        ws.append([
            json.dumps(jsonable(row.get(h)), ensure_ascii=False) if isinstance(row.get(h), (dict, list, tuple, set)) else row.get(h, "")
            for h in headers
        ])
    autofit(ws)



def write_workbook(out_path: Path, summary_rows: List[Dict[str, Any]], final_rows: List[Dict[str, Any]], raw_rows: List[Dict[str, Any]], by_market: Dict[str, List[Dict[str, Any]]], debug_rows: List[Dict[str, Any]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    data_map = {
        "summary": summary_rows,
        "final_selected": final_rows,
        "all_raw": raw_rows,
        "debug": debug_rows,
    }
    data_map.update(by_market)

    for sheet_name in SHEET_ORDER:
        rows = data_map.get(sheet_name, [])
        ws = wb.create_sheet(title=sheet_name[:31])
        append_sheet(ws, rows)

    wb.save(out_path)


# -----------------------------
# Main flow
# -----------------------------


def build_summary(target_date: str, raw_rows: List[Dict[str, Any]], final_rows: List[Dict[str, Any]], by_market: Dict[str, List[Dict[str, Any]]], page_url: str, workspace: Path) -> List[Dict[str, Any]]:
    summary: List[Dict[str, Any]] = []
    summary.append({"metric": "generated_at", "value": current_ts()})
    summary.append({"metric": "target_date", "value": target_date})
    summary.append({"metric": "page_url", "value": page_url})
    summary.append({"metric": "workspace", "value": str(workspace)})
    summary.append({"metric": "raw_total", "value": len(raw_rows)})
    summary.append({"metric": "final_total", "value": len(final_rows)})
    for market in MARKETS:
        summary.append({
            "metric": f"raw_{market.market_id}",
            "value": len(by_market.get(market.market_id, [])),
            "label": market.market_label,
        })
    return summary



def dedupe_final(raw_rows: List[Dict[str, Any]], max_picks: int) -> List[Dict[str, Any]]:
    best: Dict[str, Dict[str, Any]] = {}
    for row in raw_rows:
        key = unique_match_key(row)
        if not key:
            continue
        existing = best.get(key)
        if existing is None or safe_float(row.get("odd"), 0.0) > safe_float(existing.get("odd"), 0.0):
            best[key] = row
    final_rows = sorted(
        best.values(),
        key=lambda r: (safe_float(r.get("odd"), 0.0), safe_float(r.get("probability"), 0.0)),
        reverse=True,
    )
    return final_rows[:max_picks]



def main() -> int:
    parser = argparse.ArgumentParser(description="UI-driven Match Center dry run")
    parser.add_argument("--repo-root", default=".")
    parser.add_argument("--target-date", required=True)
    parser.add_argument("--max-picks", type=int, default=40)
    parser.add_argument("--output-dir", default="_match_center_ui_dry_run")
    parser.add_argument("--headless", action="store_true", default=True)
    args = parser.parse_args()

    repo_root = Path(args.repo_root).resolve()
    output_dir = repo_root / args.output_dir / args.target_date
    ensure_dir(output_dir)

    workspace = prepare_workspace(repo_root, args.target_date)
    debug_rows: List[Dict[str, Any]] = []
    by_market: Dict[str, List[Dict[str, Any]]] = {m.market_id: [] for m in MARKETS}

    with serve_directory(workspace) as base_url:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=args.headless)
            page = browser.new_page(viewport={"width": 1600, "height": 2200})
            page_url = open_match_center(page, base_url)
            wait_for_ui_ready(page)

            for market in MARKETS:
                try:
                    apply_market_filter(page, market)
                    extracted = extract_market_rows(page, market)
                    rows = extracted.get("rows") or []
                    by_market[market.market_id] = rows
                    debug_rows.append({
                        "market_id": market.market_id,
                        "market_label": market.market_label,
                        "heading_count": extracted.get("heading_count"),
                        "visible_count": extracted.get("visible_count"),
                        "notes": market.notes,
                    })
                except Exception as exc:
                    debug_rows.append({
                        "market_id": market.market_id,
                        "market_label": market.market_label,
                        "error": str(exc),
                        "notes": market.notes,
                    })

            browser.close()

    raw_rows: List[Dict[str, Any]] = []
    for market in MARKETS:
        raw_rows.extend(by_market.get(market.market_id, []))

    final_rows = dedupe_final(raw_rows, args.max_picks)
    summary_rows = build_summary(args.target_date, raw_rows, final_rows, by_market, page_url, workspace)

    workbook = output_dir / f"match_center_ui_dry_run_{args.target_date}_full.xlsx"
    write_workbook(workbook, summary_rows, final_rows, raw_rows, by_market, debug_rows)

    print("=" * 72)
    print("MATCH CENTER UI DRY RUN")
    print(f"Date: {args.target_date}")
    print(f"Workspace: {workspace}")
    print(f"Page: {page_url}")
    print("=" * 72)
    print(f"# Raw total: {len(raw_rows)}")
    print(f"# Final unique (cap {args.max_picks}): {len(final_rows)}")
    print("# Raw breakdown:")
    for market in MARKETS:
        print(f"- {market.market_label}: {len(by_market.get(market.market_id, []))}")
    print(f"# Workbook: {workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
