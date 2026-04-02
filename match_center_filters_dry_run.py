#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Dry run separato per selezioni Match Center basate su cache di match_predictor.

Obiettivo:
- NON tocca scraper / Supabase.
- Legge la cache giornaliera già generata da match_predictor.
- Applica regole custom sui mercati.
- Deduplica le partite: se una fixture esce in più regole, tiene la selezione con quota più alta.
- Produce output leggibile (json/csv/xlsx/md) e può girare su GitHub Actions.

Pensato per essere robusto a più shape del JSON cache:
- records con campi flat (odd_..., prob_...)
- records con array mercati/opzioni annidati
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover
    ZoneInfo = None  # type: ignore

try:
    from openpyxl import Workbook
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore


DUBLIN_TZ = ZoneInfo("Europe/Dublin") if ZoneInfo else None


RULE_ORDER = [
    "OVER_2_5",
    "OVER_1_5",
    "OVER_3_5",
    "OVER_1_5_HT",
    "DOUBLE_CHANCE_1X",
    "DOUBLE_CHANCE_X2",
    "BTTS_YES",
    "BTTS_NO",
    "CORNERS_OVER_8_5",
    "CORNERS_UNDER_10_5",
    "CARDS_OVER_4",
    "COMBO_O25_BTTS_YES",
]
RULE_PRIORITY = {rule: idx for idx, rule in enumerate(RULE_ORDER)}


@dataclass(frozen=True)
class RuleSpec:
    code: str
    label: str
    min_probability: float
    min_odd: Optional[float]


RULES: Dict[str, RuleSpec] = {
    "OVER_2_5": RuleSpec("OVER_2_5", "Over 2.5", 80.0, None),
    "OVER_1_5": RuleSpec("OVER_1_5", "Over 1.5", 80.0, 1.30),
    "OVER_3_5": RuleSpec("OVER_3_5", "Over 3.5", 75.0, None),
    "OVER_1_5_HT": RuleSpec("OVER_1_5_HT", "Over 1.5 HT", 70.0, None),
    "DOUBLE_CHANCE_1X": RuleSpec("DOUBLE_CHANCE_1X", "1X", 70.0, 1.30),
    "DOUBLE_CHANCE_X2": RuleSpec("DOUBLE_CHANCE_X2", "X2", 70.0, 1.30),
    "BTTS_YES": RuleSpec("BTTS_YES", "BTTS YES", 80.0, None),
    "BTTS_NO": RuleSpec("BTTS_NO", "BTTS NO", 80.0, None),
    "CORNERS_OVER_8_5": RuleSpec("CORNERS_OVER_8_5", "Corner Over 8.5", 80.0, None),
    "CORNERS_UNDER_10_5": RuleSpec("CORNERS_UNDER_10_5", "Corner Under 10.5", 80.0, None),
    "CARDS_OVER_4": RuleSpec("CARDS_OVER_4", "Over 4 Cards", 80.0, None),
    "COMBO_O25_BTTS_YES": RuleSpec("COMBO_O25_BTTS_YES", "Combo Over 2.5 + BTTS Yes", 80.0, None),
}


FLAT_ALIASES: Dict[str, Dict[str, List[str]]] = {
    "OVER_2_5": {
        "prob": [
            "prob_ou_2_5_over", "probability_ou_2_5_over", "prob_over_2_5",
            "probability_over_2_5", "p_over25", "over25_probability",
        ],
        "odd": [
            "odd_ou_2_5_over", "odd_over_2_5", "odds_over_2_5", "o25_odd",
        ],
    },
    "OVER_1_5": {
        "prob": [
            "prob_ou_1_5_over", "probability_ou_1_5_over", "prob_over_1_5",
            "probability_over_1_5", "p_over15", "over15_probability",
        ],
        "odd": [
            "odd_ou_1_5_over", "odd_over_1_5", "odds_over_1_5", "o15_odd",
        ],
    },
    "OVER_3_5": {
        "prob": [
            "prob_ou_3_5_over", "probability_ou_3_5_over", "prob_over_3_5",
            "probability_over_3_5", "p_over35", "over35_probability",
        ],
        "odd": [
            "odd_ou_3_5_over", "odd_over_3_5", "odds_over_3_5", "o35_odd",
        ],
    },
    "OVER_1_5_HT": {
        "prob": [
            "prob_1h_over_1_5", "probability_1h_over_1_5", "prob_over_1_5_ht",
            "probability_over_1_5_ht", "prob_ht_1_5_over",
        ],
        "odd": [
            "odd_1h_over_1_5", "odds_1h_over_1_5", "odd_over_1_5_ht",
            "odd_ht_over_1_5",
        ],
    },
    "DOUBLE_CHANCE_1X": {
        "prob": ["prob_1x", "probability_1x", "prob_double_chance_1x"],
        "odd": ["odd_1x", "odds_1x", "odd_double_chance_1x"],
    },
    "DOUBLE_CHANCE_X2": {
        "prob": ["prob_x2", "probability_x2", "prob_double_chance_x2"],
        "odd": ["odd_x2", "odds_x2", "odd_double_chance_x2"],
    },
    "BTTS_YES": {
        "prob": ["prob_btts_yes", "probability_btts_yes", "p_btts_yes", "p_btts"],
        "odd": ["odd_btts_yes", "odds_btts_yes"],
    },
    "BTTS_NO": {
        "prob": ["prob_btts_no", "probability_btts_no", "p_btts_no"],
        "odd": ["odd_btts_no", "odds_btts_no"],
    },
    "CORNERS_OVER_8_5": {
        "prob": ["prob_corner_over_8_5", "probability_corner_over_8_5"],
        "odd": ["odd_corner_over_8_5", "odds_corner_over_8_5"],
    },
    "CORNERS_UNDER_10_5": {
        "prob": ["prob_corner_under_10_5", "probability_corner_under_10_5"],
        "odd": ["odd_corner_under_10_5", "odds_corner_under_10_5"],
    },
    "CARDS_OVER_4": {
        "prob": [
            "prob_cards_over_4", "probability_cards_over_4", "prob_cards_over_4_5",
            "probability_cards_over_4_5", "prob_yellow_over_4_5",
        ],
        "odd": [
            "odd_cards_over_4", "odds_cards_over_4", "odd_cards_over_4_5",
            "odds_cards_over_4_5", "odd_yellow_over_4_5",
        ],
    },
    "COMBO_O25_BTTS_YES": {
        "prob": [
            "prob_combo_over_2_5_btts_yes", "probability_combo_over_2_5_btts_yes",
            "prob_o25_btts_yes", "prob_combo_o25_gg",
        ],
        "odd": [
            "odd_combo_over_2_5_btts_yes", "odds_combo_over_2_5_btts_yes",
            "odd_o25_btts_yes", "odd_combo_o25_gg",
        ],
    },
}


def slug(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("&", " and ")
    text = text.replace("+", " plus ")
    text = text.replace("%", " percent ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    text = text.replace("%", "")
    try:
        return float(text)
    except Exception:
        return None


def parse_probability(value: Any) -> Optional[float]:
    num = parse_float(value)
    if num is None:
        return None
    if 0.0 <= num <= 1.0:
        return round(num * 100.0, 4)
    return round(num, 4)


def first_non_empty(*values: Any) -> Optional[Any]:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return value
    return None


def get_nested(obj: Any, *path: Any) -> Any:
    cur = obj
    for part in path:
        if cur is None:
            return None
        if isinstance(cur, dict):
            cur = cur.get(part)
        elif isinstance(cur, list) and isinstance(part, int) and 0 <= part < len(cur):
            cur = cur[part]
        else:
            return None
    return cur


def tomorrow_dublin_str() -> str:
    now = datetime.now(DUBLIN_TZ) if DUBLIN_TZ else datetime.utcnow()
    return (now + timedelta(days=1)).strftime("%Y-%m-%d")


def looks_like_match_node(node: Dict[str, Any]) -> bool:
    if not isinstance(node, dict):
        return False
    direct_team_markers = (
        ("home_team" in node and "away_team" in node)
        or ("home" in node and "away" in node)
        or ("teams" in node)
    )
    nested_team_markers = bool(get_nested(node, "teams", "home")) and bool(get_nested(node, "teams", "away"))
    has_team_info = direct_team_markers or nested_team_markers
    has_fixture_marker = any(key in node for key in ("fixture_id", "fixture", "match_id", "event_id"))
    has_market_marker = (
        any(str(k).startswith("odd_") or str(k).startswith("prob") for k in node.keys())
        or any(key in node for key in ("markets", "predictions", "market_cards", "cards"))
    )
    return has_team_info and (has_fixture_marker or has_market_marker)


def collect_match_nodes(obj: Any) -> List[Dict[str, Any]]:
    found: List[Dict[str, Any]] = []

    def walk(node: Any) -> None:
        if isinstance(node, dict):
            if looks_like_match_node(node):
                found.append(node)
            for value in node.values():
                walk(value)
        elif isinstance(node, list):
            for item in node:
                walk(item)

    walk(obj)
    # dedupe by object id to avoid recursive duplicates
    uniq: List[Dict[str, Any]] = []
    seen_ids = set()
    for node in found:
        marker = id(node)
        if marker in seen_ids:
            continue
        seen_ids.add(marker)
        uniq.append(node)
    return uniq


def flatten_scalars(obj: Any, prefix: str = "") -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    if isinstance(obj, dict):
        for key, value in obj.items():
            next_prefix = f"{prefix}_{slug(key)}" if prefix else slug(key)
            out.update(flatten_scalars(value, next_prefix))
    elif isinstance(obj, list):
        for idx, value in enumerate(obj):
            next_prefix = f"{prefix}_{idx}" if prefix else str(idx)
            out.update(flatten_scalars(value, next_prefix))
    else:
        out[prefix] = obj
    return out


def choose_flat_value(flat: Dict[str, Any], aliases: Iterable[str], probability: bool = False) -> Optional[float]:
    alias_slugs = [slug(alias) for alias in aliases]
    best: Optional[float] = None
    for key, value in flat.items():
        key_slug = slug(key)
        if not any(key_slug.endswith(alias) or alias in key_slug for alias in alias_slugs):
            continue
        parsed = parse_probability(value) if probability else parse_float(value)
        if parsed is None:
            continue
        if best is None:
            best = parsed
            continue
        if probability:
            if parsed > best:
                best = parsed
        else:
            # per le quote scegliamo la più alta
            if parsed > best:
                best = parsed
    return best


def normalize_market_rule(market_text: str, pick_text: str) -> Optional[Tuple[str, str]]:
    text = f"{market_text} {pick_text}".lower().strip()
    text = text.replace("both teams score", "btts")
    text = text.replace("both teams to score", "btts")
    text = text.replace("goal goal", "btts yes")
    text = text.replace("gg", "btts yes")
    text = re.sub(r"\s+", " ", text)

    is_corner = "corner" in text
    is_card = any(token in text for token in ("card", "booking", "yellow"))
    is_combo = any(token in text for token in ("combo", "same game", "builder"))
    is_first_half = any(token in text for token in ("1st half", "first half", "1h", "ht", "primo tempo"))

    if is_combo and ("over 2.5" in text or "over 2,5" in text) and ("btts yes" in text or ("btts" in text and "yes" in text)):
        return ("COMBO_O25_BTTS_YES", "Combo Over 2.5 + BTTS Yes")

    if ("btts" in text or "both teams" in text) and "no" in text:
        return ("BTTS_NO", "BTTS NO")
    if ("btts" in text or "both teams" in text) and any(tok in text for tok in ("yes", "si", "sì")):
        return ("BTTS_YES", "BTTS YES")

    if any(token in text for token in ("double chance", "doppia chance")) or re.search(r"(^|\s)1x(\s|$)", text) or re.search(r"(^|\s)x2(\s|$)", text):
        if re.search(r"(^|\s)1x(\s|$)", text):
            return ("DOUBLE_CHANCE_1X", "1X")
        if re.search(r"(^|\s)x2(\s|$)", text):
            return ("DOUBLE_CHANCE_X2", "X2")

    if is_corner and any(token in text for token in ("under 10.5", "under 10,5", "u10.5", "u10_5")):
        return ("CORNERS_UNDER_10_5", "Corner Under 10.5")
    if is_corner and any(token in text for token in ("over 8.5", "over 8,5", "o8.5", "o8_5")):
        return ("CORNERS_OVER_8_5", "Corner Over 8.5")

    if is_card and any(token in text for token in ("over 4.5", "over 4,5", "over 4 cards", "over 4", "o4.5")):
        return ("CARDS_OVER_4", "Over 4 Cards")

    if not is_corner and not is_card and not is_combo and is_first_half and any(token in text for token in ("over 1.5", "over 1,5", "o1.5", "o1_5")):
        return ("OVER_1_5_HT", "Over 1.5 HT")

    if not is_corner and not is_card and not is_combo and any(token in text for token in ("over 3.5", "over 3,5", "o3.5", "o3_5")):
        return ("OVER_3_5", "Over 3.5")
    if not is_corner and not is_card and not is_combo and any(token in text for token in ("over 2.5", "over 2,5", "o2.5", "o2_5")):
        return ("OVER_2_5", "Over 2.5")
    if not is_corner and not is_card and not is_combo and not is_first_half and any(token in text for token in ("over 1.5", "over 1,5", "o1.5", "o1_5")):
        return ("OVER_1_5", "Over 1.5")

    return None


def extract_candidates_from_market_dict(node: Dict[str, Any], parent_market: str = "") -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    market_label = str(first_non_empty(
        node.get("market"), node.get("market_name"), node.get("marketLabel"),
        node.get("title"), node.get("group"), node.get("name"), parent_market,
    ) or "")

    options = first_non_empty(node.get("options"), node.get("outcomes"), node.get("values"), node.get("selections"))
    if isinstance(options, list) and options:
        for option in options:
            if not isinstance(option, dict):
                continue
            pick_label = str(first_non_empty(
                option.get("pick"), option.get("outcome"), option.get("selection"),
                option.get("label"), option.get("name"), option.get("value"),
            ) or "")
            odd = first_non_empty(option.get("odd"), option.get("odds"), option.get("price"), option.get("quote"))
            probability = first_non_empty(
                option.get("probability"), option.get("prob"), option.get("percent"), option.get("confidence")
            )
            norm = normalize_market_rule(market_label, pick_label)
            if norm is None:
                continue
            odd_num = parse_float(odd)
            prob_num = parse_probability(probability)
            if odd_num is None or prob_num is None:
                continue
            out.append({
                "rule_code": norm[0],
                "rule_label": norm[1],
                "market_label": market_label,
                "pick_label": pick_label,
                "odd": odd_num,
                "probability": prob_num,
                "source": "markets-array",
            })
        if out:
            return out

    pick_label = str(first_non_empty(node.get("pick"), node.get("selection"), node.get("outcome"), node.get("label"), node.get("value")) or "")
    odd = first_non_empty(node.get("odd"), node.get("odds"), node.get("price"), node.get("quote"))
    probability = first_non_empty(node.get("probability"), node.get("prob"), node.get("percent"), node.get("confidence"))
    norm = normalize_market_rule(market_label, pick_label)
    if norm is None:
        return []
    odd_num = parse_float(odd)
    prob_num = parse_probability(probability)
    if odd_num is None or prob_num is None:
        return []
    return [{
        "rule_code": norm[0],
        "rule_label": norm[1],
        "market_label": market_label,
        "pick_label": pick_label,
        "odd": odd_num,
        "probability": prob_num,
        "source": "market-node",
    }]


def extract_market_candidates(record: Dict[str, Any]) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []

    def walk(node: Any, parent_market: str = "") -> None:
        if isinstance(node, dict):
            market_name = str(first_non_empty(node.get("market"), node.get("market_name"), node.get("title"), parent_market) or "")
            extracted = extract_candidates_from_market_dict(node, parent_market=parent_market)
            if extracted:
                candidates.extend(extracted)
            for value in node.values():
                walk(value, parent_market=market_name or parent_market)
        elif isinstance(node, list):
            for item in node:
                walk(item, parent_market=parent_market)

    walk(record)

    flat = flatten_scalars(record)
    for rule_code, aliases in FLAT_ALIASES.items():
        prob = choose_flat_value(flat, aliases.get("prob", []), probability=True)
        odd = choose_flat_value(flat, aliases.get("odd", []), probability=False)
        if prob is None or odd is None:
            continue
        rule = RULES[rule_code]
        candidates.append({
            "rule_code": rule.code,
            "rule_label": rule.label,
            "market_label": rule.label,
            "pick_label": rule.label,
            "odd": odd,
            "probability": prob,
            "source": "flat-fields",
        })

    # dedupe same rule inside same record: keep higher odd, then higher probability
    best_by_rule: Dict[str, Dict[str, Any]] = {}
    for cand in candidates:
        code = cand["rule_code"]
        prev = best_by_rule.get(code)
        if prev is None:
            best_by_rule[code] = cand
            continue
        prev_key = (float(prev["odd"]), float(prev["probability"]))
        new_key = (float(cand["odd"]), float(cand["probability"]))
        if new_key > prev_key:
            best_by_rule[code] = cand
    return list(best_by_rule.values())


def normalize_match(record: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    fixture_id = first_non_empty(
        record.get("fixture_id"), record.get("match_id"), record.get("event_id"),
        get_nested(record, "fixture", "id"), get_nested(record, "fixture", "fixture_id"),
    )

    home_name = first_non_empty(
        record.get("home_team"), record.get("home"), get_nested(record, "home", "name"),
        get_nested(record, "home", "team_name"), get_nested(record, "teams", "home", "name"),
    )
    away_name = first_non_empty(
        record.get("away_team"), record.get("away"), get_nested(record, "away", "name"),
        get_nested(record, "away", "team_name"), get_nested(record, "teams", "away", "name"),
    )
    if isinstance(home_name, dict):
        home_name = first_non_empty(home_name.get("name"), home_name.get("team_name"))
    if isinstance(away_name, dict):
        away_name = first_non_empty(away_name.get("name"), away_name.get("team_name"))

    home_name = str(home_name or "").strip()
    away_name = str(away_name or "").strip()
    if not home_name or not away_name:
        return None

    league_name = first_non_empty(
        record.get("league_name"), record.get("league"), get_nested(record, "league", "name")
    )
    country = first_non_empty(record.get("country"), get_nested(record, "league", "country"))
    date_value = first_non_empty(record.get("date"), get_nested(record, "fixture", "date"), record.get("match_date"))
    time_value = first_non_empty(record.get("time"), get_nested(record, "fixture", "time"), record.get("kickoff"))

    match_key = first_non_empty(record.get("match_key"), record.get("fixture_key"))
    if not match_key:
        match_key = f"{fixture_id or ''}|{home_name}|{away_name}|{date_value or ''}|{time_value or ''}"

    return {
        "fixture_id": str(fixture_id or match_key).strip(),
        "match_key": str(match_key).strip(),
        "date": str(date_value or "").strip(),
        "time": str(time_value or "").strip(),
        "league": str(league_name or "").strip(),
        "country": str(country or "").strip(),
        "home": home_name,
        "away": away_name,
        "raw": record,
    }


def load_cache_json(cache_file: Path) -> Any:
    with cache_file.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def resolve_cache_file(repo_root: Path, cache_dir: str, target_date: str, explicit_file: Optional[str]) -> Path:
    if explicit_file:
        path = Path(explicit_file)
        if not path.is_absolute():
            path = repo_root / path
        if not path.exists():
            raise FileNotFoundError(f"Cache file non trovato: {path}")
        return path

    base = repo_root / cache_dir
    dated = base / f"{target_date}.json"
    if dated.exists():
        return dated
    latest = base / "latest.json"
    if latest.exists():
        return latest
    raise FileNotFoundError(f"Nessun cache file trovato in {base} per {target_date} (né latest.json)")


def apply_rules(matches: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, int], Dict[str, int]]:
    all_candidates: List[Dict[str, Any]] = []
    pre_rule_counter: Counter[str] = Counter()

    for match in matches:
        candidates = extract_market_candidates(match["raw"])
        for cand in candidates:
            spec = RULES.get(cand["rule_code"])
            if spec is None:
                continue
            prob = float(cand["probability"])
            odd = float(cand["odd"])
            if prob < spec.min_probability:
                continue
            if spec.min_odd is not None and odd < spec.min_odd:
                continue
            row = {
                "fixture_id": match["fixture_id"],
                "match_key": match["match_key"],
                "date": match["date"],
                "time": match["time"],
                "league": match["league"],
                "country": match["country"],
                "home": match["home"],
                "away": match["away"],
                "rule_code": spec.code,
                "rule_label": spec.label,
                "market_label": cand["market_label"],
                "pick_label": cand["pick_label"],
                "probability": round(prob, 2),
                "odd": round(odd, 3),
                "source": cand["source"],
            }
            all_candidates.append(row)
            pre_rule_counter[spec.code] += 1

    best_by_match: Dict[str, Dict[str, Any]] = {}
    for row in all_candidates:
        key = row["match_key"] or row["fixture_id"]
        prev = best_by_match.get(key)
        if prev is None:
            best_by_match[key] = row
            continue
        prev_cmp = (
            float(prev["odd"]),
            float(prev["probability"]),
            -RULE_PRIORITY.get(prev["rule_code"], 999),
        )
        new_cmp = (
            float(row["odd"]),
            float(row["probability"]),
            -RULE_PRIORITY.get(row["rule_code"], 999),
        )
        if new_cmp > prev_cmp:
            best_by_match[key] = row

    selected = list(best_by_match.values())
    selected.sort(key=lambda r: (float(r["probability"]), float(r["odd"]), -RULE_PRIORITY.get(r["rule_code"], 999)), reverse=True)

    post_rule_counter: Counter[str] = Counter(row["rule_code"] for row in selected)
    return all_candidates, selected, dict(pre_rule_counter), dict(post_rule_counter)


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    fieldnames = [
        "fixture_id", "match_key", "date", "time", "country", "league",
        "home", "away", "rule_code", "rule_label", "market_label", "pick_label",
        "probability", "odd", "source",
    ]
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def write_xlsx(path: Path, selected: List[Dict[str, Any]], all_candidates: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    if Workbook is None:
        return
    wb = Workbook()
    ws_sel = wb.active
    ws_sel.title = "selected"
    headers = [
        "fixture_id", "date", "time", "country", "league", "home", "away",
        "rule_label", "probability", "odd", "market_label", "pick_label", "source",
    ]
    ws_sel.append(headers)
    for row in selected:
        ws_sel.append([row.get(h, "") for h in headers])

    ws_all = wb.create_sheet("all_candidates")
    ws_all.append(headers)
    for row in all_candidates:
        ws_all.append([row.get(h, "") for h in headers])

    ws_sum = wb.create_sheet("summary")
    ws_sum.append(["key", "value"])
    for key, value in summary.items():
        if isinstance(value, (dict, list)):
            ws_sum.append([key, json.dumps(value, ensure_ascii=False)])
        else:
            ws_sum.append([key, value])

    wb.save(path)


def write_markdown(path: Path, target_date: str, cache_file: Path, selected: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    lines = []
    lines.append(f"# Dry run Match Center — {target_date}")
    lines.append("")
    lines.append(f"- Cache usata: `{cache_file}`")
    lines.append(f"- Match letti: **{summary['matches_found']}**")
    lines.append(f"- Candidati che passano i filtri: **{summary['all_candidates_count']}**")
    lines.append(f"- Selezioni univoche prima del cap: **{summary['selected_before_cap']}**")
    lines.append(f"- Selezioni finali (cap {summary['max_picks']}): **{summary['selected_after_cap']}**")
    lines.append("")
    lines.append("## Breakdown per regola (prima della deduplica)")
    lines.append("")
    for rule_code in RULE_ORDER:
        lines.append(f"- {RULES[rule_code].label}: {summary['pre_rule_counts'].get(rule_code, 0)}")
    lines.append("")
    lines.append("## Breakdown per regola (dopo deduplica/cap)")
    lines.append("")
    for rule_code in RULE_ORDER:
        lines.append(f"- {RULES[rule_code].label}: {summary['post_rule_counts'].get(rule_code, 0)}")
    lines.append("")
    lines.append("## Selezioni finali")
    lines.append("")
    lines.append("| # | Match | Regola | Prob. | Quota | Lega |")
    lines.append("|---:|---|---|---:|---:|---|")
    for idx, row in enumerate(selected, start=1):
        match_name = f"{row['home']} vs {row['away']}"
        lines.append(
            f"| {idx} | {match_name} | {row['rule_label']} | {row['probability']:.2f}% | {row['odd']:.2f} | {row['league']} |"
        )
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Dry run Match Center filters from match_predictor cache")
    parser.add_argument("--repo-root", default=".", help="Root repo path")
    parser.add_argument("--cache-dir", default="assets/data/match-predictor", help="Cache dir relativo al repo")
    parser.add_argument("--cache-file", default="", help="Cache JSON esplicito")
    parser.add_argument("--date", default="", help="Data target YYYY-MM-DD. Default: domani Europe/Dublin")
    parser.add_argument("--max-picks", type=int, default=40, help="Cap finale selezioni")
    parser.add_argument("--output-dir", default="_match_center_filter_dry_run", help="Cartella output")
    args = parser.parse_args()

    target_date = args.date.strip() or tomorrow_dublin_str()
    repo_root = Path(args.repo_root).resolve()
    output_dir = (repo_root / args.output_dir / target_date).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    cache_file = resolve_cache_file(repo_root, args.cache_dir, target_date, args.cache_file.strip() or None)
    payload = load_cache_json(cache_file)
    raw_matches = collect_match_nodes(payload)

    matches: List[Dict[str, Any]] = []
    seen_match_keys = set()
    for node in raw_matches:
        normalized = normalize_match(node)
        if not normalized:
            continue
        match_key = normalized["match_key"]
        if match_key in seen_match_keys:
            continue
        seen_match_keys.add(match_key)
        matches.append(normalized)

    all_candidates, selected_unique, pre_rule_counts, post_rule_counts = apply_rules(matches)
    selected_capped = selected_unique[: max(0, args.max_picks)]

    summary = {
        "target_date": target_date,
        "cache_file": str(cache_file),
        "matches_found": len(matches),
        "all_candidates_count": len(all_candidates),
        "selected_before_cap": len(selected_unique),
        "selected_after_cap": len(selected_capped),
        "max_picks": int(args.max_picks),
        "pre_rule_counts": pre_rule_counts,
        "post_rule_counts": dict(Counter(row["rule_code"] for row in selected_capped)),
    }

    write_json(output_dir / f"match_center_dry_run_{target_date}_summary.json", summary)
    write_json(output_dir / f"match_center_dry_run_{target_date}_all_candidates.json", all_candidates)
    write_json(output_dir / f"match_center_dry_run_{target_date}_selected.json", selected_capped)
    write_csv(output_dir / f"match_center_dry_run_{target_date}_all_candidates.csv", all_candidates)
    write_csv(output_dir / f"match_center_dry_run_{target_date}_selected.csv", selected_capped)
    write_markdown(output_dir / f"match_center_dry_run_{target_date}_report.md", target_date, cache_file, selected_capped, summary)
    write_xlsx(output_dir / f"match_center_dry_run_{target_date}.xlsx", selected_capped, all_candidates, summary)

    print("=" * 72)
    print("MATCH CENTER DRY RUN")
    print(f"Date: {target_date}")
    print(f"Cache: {cache_file}")
    print("=" * 72)
    print(f"# Matches normalized: {len(matches)}")
    print(f"# All candidates passing filters: {len(all_candidates)}")
    print(f"# Selected unique before cap: {len(selected_unique)}")
    print(f"# Final selected (cap {args.max_picks}): {len(selected_capped)}")
    print("")
    print("# Breakdown after cap:")
    for rule_code in RULE_ORDER:
        print(f"- {RULES[rule_code].label}: {summary['post_rule_counts'].get(rule_code, 0)}")
    print("")
    for idx, row in enumerate(selected_capped, start=1):
        print(
            f"{idx:02d}. {row['home']} vs {row['away']} | {row['rule_label']} | "
            f"{row['probability']:.2f}% | {row['odd']:.2f} | {row['league']}"
        )
    print("")
    print(f"# Output dir: {output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
