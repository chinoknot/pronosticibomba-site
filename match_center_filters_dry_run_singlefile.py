#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None


MARKET_RULES = [
    {"id": "over_25", "label": "Over 2.5", "min_prob": 80.0, "min_odd": None},
    {"id": "over_15", "label": "Over 1.5", "min_prob": 80.0, "min_odd": 1.30},
    {"id": "over_35", "label": "Over 3.5", "min_prob": 75.0, "min_odd": None},
    {"id": "over_15_ht", "label": "Over 1.5 HT", "min_prob": 70.0, "min_odd": None},
    {"id": "dc_1x", "label": "1X", "min_prob": 70.0, "min_odd": 1.30},
    {"id": "dc_x2", "label": "X2", "min_prob": 70.0, "min_odd": 1.30},
    {"id": "btts_yes", "label": "BTTS YES", "min_prob": 80.0, "min_odd": None},
    {"id": "btts_no", "label": "BTTS NO", "min_prob": 80.0, "min_odd": None},
    {"id": "corner_over_85", "label": "Corner Over 8.5", "min_prob": 80.0, "min_odd": None},
    {"id": "corner_under_105", "label": "Corner Under 10.5", "min_prob": 80.0, "min_odd": None},
    {"id": "cards_over_4", "label": "Over 4 Cards", "min_prob": 80.0, "min_odd": None},
    {"id": "combo_o25_btts_yes", "label": "Combo Over 2.5 + BTTS Yes", "min_prob": 80.0, "min_odd": None},
]
RULE_BY_ID = {r["id"]: r for r in MARKET_RULES}


def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("%", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def prob_to_percent(value: Any) -> Optional[float]:
    f = to_float(value)
    if f is None:
        return None
    if 0 <= f <= 1.0:
        return f * 100.0
    if 0 <= f <= 100.0:
        return f
    return None


def normalize_text(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(text or "").lower()).strip()


def compact_text(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text or "").lower())


def implied_prob_from_odd(odd: Optional[float]) -> Optional[float]:
    if odd is None or odd <= 1.0:
        return None
    return 100.0 / odd


def poisson_over_prob(lam: Optional[float], line: float) -> Optional[float]:
    if lam is None or lam <= 0:
        return None
    threshold = int(math.floor(line) + 1)
    p_le = 0.0
    for k in range(threshold):
        p_le += math.exp(-lam) * (lam ** k) / math.factorial(k)
    return max(0.0, min(100.0, (1.0 - p_le) * 100.0))


def poisson_under_prob(lam: Optional[float], line: float) -> Optional[float]:
    p_over = poisson_over_prob(lam, line)
    if p_over is None:
        return None
    return max(0.0, min(100.0, 100.0 - p_over))


def poisson_btts_yes(lam_home: Optional[float], lam_away: Optional[float]) -> Optional[float]:
    if lam_home is None or lam_away is None or lam_home < 0 or lam_away < 0:
        return None
    p_h0 = math.exp(-lam_home)
    p_a0 = math.exp(-lam_away)
    p_yes = 1.0 - p_h0 - p_a0 + (p_h0 * p_a0)
    return max(0.0, min(100.0, p_yes * 100.0))


def flatten(obj: Any, prefix: str = "") -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            key = f"{prefix}.{k}" if prefix else str(k)
            out.update(flatten(v, key))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            key = f"{prefix}[{i}]"
            out.update(flatten(v, key))
    else:
        out[prefix] = obj
    return out


def nested_get(data: Any, path: str) -> Any:
    cur = data
    for part in path.split("."):
        if not isinstance(cur, dict):
            return None
        cur = cur.get(part)
    return cur


def find_number_by_tokens(flat: Dict[str, Any], token_sets: Iterable[Tuple[str, ...]], prob: bool = False) -> Optional[float]:
    for key, value in flat.items():
        nk = compact_text(key)
        if not nk:
            continue
        for tokens in token_sets:
            if all(tok in nk for tok in tokens):
                found = prob_to_percent(value) if prob else to_float(value)
                if found is not None:
                    return found
    return None


def parse_dt(value: Any) -> str:
    s = str(value or "").strip()
    if not s:
        return ""
    try:
        if len(s) >= 16 and "T" in s:
            return s[11:16]
    except Exception:
        pass
    return s[:5] if len(s) >= 5 else s

def find_text_by_tokens(flat: Dict[str, Any], token_sets: Iterable[Tuple[str, ...]]) -> str:
    for key, value in flat.items():
        nk = compact_text(key)
        if not nk:
            continue
        for tokens in token_sets:
            if all(tok in nk for tok in tokens):
                if value is None:
                    continue
                s = str(value).strip()
                if s:
                    return s
    return ""


def find_time_candidate(flat: Dict[str, Any]) -> str:
    # direct time-like keys first
    for key, value in flat.items():
        nk = compact_text(key)
        if any(tok in nk for tok in ('time','kickoff','start','matchtime')):
            s = parse_dt(value)
            if s:
                return s
    # ISO datetime anywhere
    for value in flat.values():
        s = parse_dt(value)
        if s and len(s) == 5 and s[2] == ':':
            return s
    return ""


def find_date_candidate(flat: Dict[str, Any]) -> str:
    for key, value in flat.items():
        nk = compact_text(key)
        if any(tok in nk for tok in ('date','fixturedate','kickoff','start')):
            s = str(value or '').strip()
            if len(s) >= 10 and s[4] == '-' and s[7] == '-':
                return s[:10]
    for value in flat.values():
        s = str(value or '').strip()
        if len(s) >= 10 and s[4] == '-' and s[7] == '-':
            return s[:10]
    return ""


def resolve_cache_file(repo_root: Path, cache_dir: str, target_date: str, explicit: Optional[str]) -> Path:
    if explicit:
        path = Path(explicit)
        if not path.is_absolute():
            path = repo_root / path
        if not path.exists():
            raise FileNotFoundError(f"Cache file esplicito non trovato: {path}")
        return path
    base = repo_root / cache_dir
    dated = base / f"{target_date}.json"
    if dated.exists():
        return dated
    latest = base / "latest.json"
    if latest.exists():
        return latest
    raise FileNotFoundError(f"Nessun cache file trovato in {base} per {target_date} (né latest.json)")


def iter_match_candidates(root: Any) -> Iterable[dict]:
    if isinstance(root, list):
        for item in root:
            if isinstance(item, dict):
                yield item
        return
    if not isinstance(root, dict):
        return
    # common containers
    for key in ("matches", "fixtures", "data", "items", "rows"):
        val = root.get(key)
        if isinstance(val, list):
            for item in val:
                if isinstance(item, dict):
                    yield item
            return
    # manifest-like / date keyed
    for value in root.values():
        if isinstance(value, list):
            if value and all(isinstance(x, dict) for x in value):
                for item in value:
                    yield item
                return
        if isinstance(value, dict):
            sub = value.get("matches") or value.get("fixtures") or value.get("data")
            if isinstance(sub, list) and sub and all(isinstance(x, dict) for x in sub):
                for item in sub:
                    yield item
                return
    # fallback: dict itself may be a match map
    if any(k in root for k in ("fixture_id", "home_team", "away_team", "teams", "fixture")):
        yield root


def canonical_market(label: str) -> Optional[str]:
    n = normalize_text(label)
    c = compact_text(label)

    if ("combo" in n or "same game" in n or "sgp" in n) and (("over 2 5" in n) or ("over25" in c)) and (("btts" in c) or ("both teams" in n)) and ("yes" in n):
        return "combo_o25_btts_yes"
    if ("corner" in n or "corners" in n) and (("under 10 5" in n) or ("under105" in c)):
        return "corner_under_105"
    if ("corner" in n or "corners" in n) and (("over 8 5" in n) or ("over85" in c)):
        return "corner_over_85"
    if (("card" in n) or ("cards" in n) or ("yellow" in n)) and (("over 4" in n) or ("over4" in c)):
        return "cards_over_4"
    if (("first half" in n) or ("1st half" in n) or ("ht" in n)) and (("over 1 5" in n) or ("over15" in c)):
        return "over_15_ht"
    if ("double chance" in n and "1x" in c) or (c == "1x"):
        return "dc_1x"
    if ("double chance" in n and "x2" in c) or (c == "x2"):
        return "dc_x2"
    if (("btts" in c) or ("both teams" in n)) and ("yes" in n):
        return "btts_yes"
    if (("btts" in c) or ("both teams" in n)) and ("no" in n):
        return "btts_no"
    if (("over 3 5" in n) or ("over35" in c)) and "half" not in n:
        return "over_35"
    if (("over 2 5" in n) or ("over25" in c)) and "half" not in n and "combo" not in n:
        return "over_25"
    if (("over 1 5" in n) or ("over15" in c)) and "half" not in n and "combo" not in n:
        return "over_15"
    return None


def maybe_market_objects(node: Any, collector: List[dict], hint: str = "") -> None:
    if isinstance(node, dict):
        nameish = [node.get(k) for k in ("name", "label", "market", "key", "slug", "title", "bet", "selection") if node.get(k) is not None]
        outcomeish = [node.get(k) for k in ("outcome", "option", "pick", "side", "value") if node.get(k) is not None]
        label = " ".join(str(x) for x in [hint] + nameish + outcomeish if str(x).strip())
        market_id = canonical_market(label)
        prob = None
        odd = None
        for k in ("probability", "prob", "chance", "confidence", "pct", "percent"):
            if k in node:
                prob = prob_to_percent(node.get(k))
                if prob is not None:
                    break
        for k in ("odd", "odds", "price", "quote"):
            if k in node:
                odd = to_float(node.get(k))
                if odd is not None:
                    break
        if market_id and (prob is not None or odd is not None):
            collector.append({"market_id": market_id, "probability": prob, "odd": odd, "source": label})
        for k, v in node.items():
            child_hint = f"{hint} {k}".strip() if isinstance(k, str) else hint
            maybe_market_objects(v, collector, child_hint)
    elif isinstance(node, list):
        for item in node:
            maybe_market_objects(item, collector, hint)


def extract_direct_markets(match: dict) -> Dict[str, dict]:
    found: Dict[str, dict] = {}
    candidates: List[dict] = []
    maybe_market_objects(match, candidates)
    for item in candidates:
        market_id = item["market_id"]
        prev = found.get(market_id)
        new_odd = item.get("odd") or 0.0
        prev_odd = (prev or {}).get("odd") or 0.0
        if prev is None or new_odd > prev_odd:
            found[market_id] = item
    return found


def build_fallback_markets(match: dict, flat: Dict[str, Any]) -> Dict[str, dict]:
    out: Dict[str, dict] = {}

    home_prob = find_number_by_tokens(flat, [
        ("prob", "home"), ("probability", "home"), ("homewin",), ("percent", "home")
    ], prob=True)
    draw_prob = find_number_by_tokens(flat, [
        ("prob", "draw"), ("probability", "draw"), ("drawprob",), ("percent", "draw")
    ], prob=True)
    away_prob = find_number_by_tokens(flat, [
        ("prob", "away"), ("probability", "away"), ("awaywin",), ("percent", "away")
    ], prob=True)

    odd_1x = find_number_by_tokens(flat, [
        ("odd", "1x"), ("odds", "1x"), ("doublechance", "1x")
    ])
    odd_x2 = find_number_by_tokens(flat, [
        ("odd", "x2"), ("odds", "x2"), ("doublechance", "x2")
    ])

    if home_prob is not None and draw_prob is not None:
        out["dc_1x"] = {"market_id": "dc_1x", "probability": min(100.0, home_prob + draw_prob), "odd": odd_1x, "source": "fallback_1x"}
    if away_prob is not None and draw_prob is not None:
        out["dc_x2"] = {"market_id": "dc_x2", "probability": min(100.0, away_prob + draw_prob), "odd": odd_x2, "source": "fallback_x2"}

    lam_home = find_number_by_tokens(flat, [
        ("prediction", "goals", "home"), ("expected", "goals", "home"), ("xg", "home"), ("lambda", "home")
    ])
    lam_away = find_number_by_tokens(flat, [
        ("prediction", "goals", "away"), ("expected", "goals", "away"), ("xg", "away"), ("lambda", "away")
    ])
    lam_total = find_number_by_tokens(flat, [
        ("expected", "goals", "total"), ("xg", "total"), ("lambda", "total"), ("predtotal",)
    ])
    if lam_total is None and lam_home is not None and lam_away is not None:
        lam_total = lam_home + lam_away

    odd_o15 = find_number_by_tokens(flat, [
        ("odd", "over", "15"), ("odds", "over", "15"), ("over15", "odd")
    ])
    odd_o25 = find_number_by_tokens(flat, [
        ("odd", "over", "25"), ("odds", "over", "25"), ("over25", "odd")
    ])
    odd_o35 = find_number_by_tokens(flat, [
        ("odd", "over", "35"), ("odds", "over", "35"), ("over35", "odd")
    ])

    for market_id, line, odd in (("over_15", 1.5, odd_o15), ("over_25", 2.5, odd_o25), ("over_35", 3.5, odd_o35)):
        prob = poisson_over_prob(lam_total, line)
        if prob is not None:
            out[market_id] = {"market_id": market_id, "probability": prob, "odd": odd, "source": f"fallback_poisson_{market_id}"}

    lam_ht_total = find_number_by_tokens(flat, [
        ("expected", "goals", "ht", "total"), ("xg", "ht", "total"), ("firsthalf", "goals", "total"), ("lambda", "ht")
    ])
    if lam_ht_total is None and lam_total is not None:
        lam_ht_total = lam_total * 0.46
    odd_o15_ht = find_number_by_tokens(flat, [
        ("odd", "over", "15", "ht"), ("odds", "over", "15", "ht"), ("firsthalf", "over15", "odd")
    ])
    prob_o15_ht = poisson_over_prob(lam_ht_total, 1.5)
    if prob_o15_ht is not None:
        out["over_15_ht"] = {"market_id": "over_15_ht", "probability": prob_o15_ht, "odd": odd_o15_ht, "source": "fallback_poisson_over_15_ht"}

    prob_btts_yes = find_number_by_tokens(flat, [
        ("prob", "btts", "yes"), ("probability", "btts", "yes"), ("bttsyes",)
    ], prob=True)
    prob_btts_no = find_number_by_tokens(flat, [
        ("prob", "btts", "no"), ("probability", "btts", "no"), ("bttsno",)
    ], prob=True)
    odd_btts_yes = find_number_by_tokens(flat, [
        ("odd", "btts", "yes"), ("odds", "btts", "yes"), ("bttsyes", "odd")
    ])
    odd_btts_no = find_number_by_tokens(flat, [
        ("odd", "btts", "no"), ("odds", "btts", "no"), ("bttsno", "odd")
    ])
    if prob_btts_yes is None:
        prob_btts_yes = poisson_btts_yes(lam_home, lam_away)
    if prob_btts_yes is not None:
        out["btts_yes"] = {"market_id": "btts_yes", "probability": prob_btts_yes, "odd": odd_btts_yes, "source": "fallback_btts_yes"}
        out["btts_no"] = {"market_id": "btts_no", "probability": max(0.0, 100.0 - prob_btts_yes), "odd": odd_btts_no, "source": "fallback_btts_no"}
    elif prob_btts_no is not None:
        out["btts_no"] = {"market_id": "btts_no", "probability": prob_btts_no, "odd": odd_btts_no, "source": "fallback_btts_no_direct"}
        out["btts_yes"] = {"market_id": "btts_yes", "probability": max(0.0, 100.0 - prob_btts_no), "odd": odd_btts_yes, "source": "fallback_btts_yes_direct"}

    corner_exp = find_number_by_tokens(flat, [
        ("expected", "corners", "total"), ("corners", "expected", "total"), ("corner", "total", "expected"), ("corners", "total")
    ])
    odd_co85 = find_number_by_tokens(flat, [
        ("odd", "corner", "over", "85"), ("odds", "corner", "over", "85"), ("cornerover85", "odd")
    ])
    odd_cu105 = find_number_by_tokens(flat, [
        ("odd", "corner", "under", "105"), ("odds", "corner", "under", "105"), ("cornerunder105", "odd")
    ])
    if corner_exp is not None:
        p_co85 = poisson_over_prob(corner_exp, 8.5)
        p_cu105 = poisson_under_prob(corner_exp, 10.5)
        if p_co85 is not None:
            out["corner_over_85"] = {"market_id": "corner_over_85", "probability": p_co85, "odd": odd_co85, "source": "fallback_corner_over_85"}
        if p_cu105 is not None:
            out["corner_under_105"] = {"market_id": "corner_under_105", "probability": p_cu105, "odd": odd_cu105, "source": "fallback_corner_under_105"}

    cards_exp = find_number_by_tokens(flat, [
        ("expected", "cards", "total"), ("cards", "expected", "total"), ("yellow", "total", "expected"), ("cards", "total")
    ])
    odd_cards4 = find_number_by_tokens(flat, [
        ("odd", "cards", "over", "4"), ("odds", "cards", "over", "4"), ("over4cards", "odd")
    ])
    if cards_exp is not None:
        p_cards4 = poisson_over_prob(cards_exp, 4.0)
        if p_cards4 is not None:
            out["cards_over_4"] = {"market_id": "cards_over_4", "probability": p_cards4, "odd": odd_cards4, "source": "fallback_cards_over_4"}

    if "over_25" in out and "btts_yes" in out:
        p_o25 = out["over_25"]["probability"] / 100.0
        p_btts = out["btts_yes"]["probability"] / 100.0
        # conservative blend, not pure independence
        combo_prob = min(100.0, max(0.0, ((min(p_o25, p_btts) * 0.70) + (p_o25 * p_btts * 0.30)) * 100.0))
        combo_odd = find_number_by_tokens(flat, [
            ("odd", "combo", "over", "25", "btts", "yes"),
            ("odds", "combo", "over", "25", "btts", "yes"),
            ("comboover25bttsyes", "odd"),
        ])
        out["combo_o25_btts_yes"] = {
            "market_id": "combo_o25_btts_yes",
            "probability": combo_prob,
            "odd": combo_odd,
            "source": "fallback_combo_o25_btts_yes",
        }

    return out


def normalize_match(raw: dict) -> dict:
    flat = flatten(raw)

    fixture_id = raw.get("fixture_id") or nested_get(raw, "fixture.id") or raw.get("id") or find_text_by_tokens(flat, [("fixture","id"), ("match","id"), ("game","id")])
    league = raw.get("league_name") or nested_get(raw, "league.name") or raw.get("league") or find_text_by_tokens(flat, [("league","name"), ("competition","name")])
    country = raw.get("country") or nested_get(raw, "league.country") or find_text_by_tokens(flat, [("league","country"), ("competition","country")])
    home = raw.get("home_team") or nested_get(raw, "teams.home.name") or nested_get(raw, "home.name") or find_text_by_tokens(flat, [("hometeam",), ("teamshome", "name"), ("home","name")])
    away = raw.get("away_team") or nested_get(raw, "teams.away.name") or nested_get(raw, "away.name") or find_text_by_tokens(flat, [("awayteam",), ("teamsaway", "name"), ("away","name")])
    date = raw.get("date") or nested_get(raw, "fixture.date") or find_date_candidate(flat)
    time = raw.get("time") or parse_dt(nested_get(raw, "fixture.date")) or parse_dt(raw.get("kickoff")) or find_time_candidate(flat)
    match_key = raw.get("match_key") or f"{date}|{time}|{league}|{home}|{away}"

    direct = extract_direct_markets(raw)
    fallback = build_fallback_markets(raw, flat)

    merged: Dict[str, dict] = {}
    for market_id in RULE_BY_ID:
        item = deepcopy(direct.get(market_id) or fallback.get(market_id) or {})
        if item:
            item["market_id"] = market_id
            merged[market_id] = item

    return {
        "fixture_id": str(fixture_id or "").strip(),
        "match_key": str(match_key or "").strip(),
        "date": str(date or ""),
        "time": str(time or ""),
        "league": str(league or ""),
        "country": str(country or ""),
        "home": str(home or ""),
        "away": str(away or ""),
        "markets": merged,
        "flat_keys_sample": list(flat.keys())[:120],
    }


def apply_rules(match: dict) -> Tuple[List[dict], dict]:
    candidates = []
    reasons = {rule["id"]: "missing" for rule in MARKET_RULES}
    for rule in MARKET_RULES:
        market = match["markets"].get(rule["id"])
        if not market:
            continue
        prob = prob_to_percent(market.get("probability"))
        odd = to_float(market.get("odd"))
        if prob is None:
            reasons[rule["id"]] = "no_probability"
            continue
        if prob < rule["min_prob"]:
            reasons[rule["id"]] = f"prob<{rule['min_prob']}"
            continue
        if rule["min_odd"] is not None:
            if odd is None:
                reasons[rule["id"]] = f"no_odd_min_{rule['min_odd']}"
                continue
            if odd < rule["min_odd"]:
                reasons[rule["id"]] = f"odd<{rule['min_odd']}"
                continue
        reasons[rule["id"]] = "pass"
        candidates.append({
            "fixture_id": match["fixture_id"],
            "match_key": match["match_key"],
            "date": match["date"],
            "time": match["time"],
            "league": match["league"],
            "country": match["country"],
            "home": match["home"],
            "away": match["away"],
            "market_id": rule["id"],
            "market_label": rule["label"],
            "probability": round(prob, 2),
            "odd": round(odd, 3) if odd is not None else None,
            "source": market.get("source", ""),
            "implied_probability": round(implied_prob_from_odd(odd), 2) if odd else None,
        })
    return candidates, reasons


def choose_unique_highest_odd(candidates: List[dict]) -> List[dict]:
    best: Dict[str, dict] = {}
    for item in candidates:
        key = item.get("match_key") or item.get("fixture_id")
        if not key:
            continue
        prev = best.get(key)
        new_odd = item.get("odd") if item.get("odd") is not None else -1.0
        prev_odd = prev.get("odd") if prev and prev.get("odd") is not None else -1.0
        if prev is None or new_odd > prev_odd or (new_odd == prev_odd and (item.get("probability") or 0.0) > (prev.get("probability") or 0.0)):
            best[key] = item
    selected = list(best.values())
    selected.sort(key=lambda x: (x.get("probability") or 0.0, x.get("odd") or 0.0, x.get("market_label") or ""), reverse=True)
    return selected


def write_csv(path: Path, rows: List[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = []
    seen = set()
    for row in rows:
        for k in row.keys():
            if k not in seen:
                seen.add(k)
                fieldnames.append(k)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def _append_sheet(ws, rows: List[dict]) -> None:
    if not rows:
        ws.append(["empty"])
        return
    seen = []
    seen_set = set()
    for row in rows:
        for k in row.keys():
            if k not in seen_set:
                seen.append(k)
                seen_set.add(k)
    ws.append(seen)
    for row in rows:
        ws.append([row.get(h, "") for h in seen])


def write_xlsx(path: Path, selected: List[dict], all_candidates: List[dict], summary: dict, by_market: Dict[str, List[dict]], debug_samples: List[dict]) -> None:
    if Workbook is None:
        return
    wb = Workbook()

    ws = wb.active
    ws.title = "summary"
    ws.append(["key", "value"])
    for k, v in summary.items():
        if isinstance(v, (dict, list)):
            ws.append([k, json.dumps(v, ensure_ascii=False)])
        else:
            ws.append([k, v])

    ws = wb.create_sheet("final_selected")
    _append_sheet(ws, selected)

    ws = wb.create_sheet("all_candidates")
    _append_sheet(ws, all_candidates)

    for rule in MARKET_RULES:
        sheet_name = rule["id"][:31]
        ws = wb.create_sheet(sheet_name)
        _append_sheet(ws, by_market.get(rule["id"], []))

    ws = wb.create_sheet("debug_samples")
    _append_sheet(ws, debug_samples)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_report(path: Path, target_date: str, cache_file: Path, normalized_count: int, all_candidates: List[dict], selected: List[dict]) -> None:
    counts = Counter(item["market_label"] for item in selected)
    lines = [
        "# MATCH CENTER FILTERS DRY RUN V2",
        "",
        f"- Date: {target_date}",
        f"- Cache: {cache_file}",
        f"- Matches normalized: {normalized_count}",
        f"- All candidates passing filters: {len(all_candidates)}",
        f"- Selected unique before cap: {len(choose_unique_highest_odd(all_candidates))}",
        f"- Final selected: {len(selected)}",
        "",
        "## Breakdown after cap",
    ]
    for rule in MARKET_RULES:
        lines.append(f"- {rule['label']}: {counts.get(rule['label'], 0)}")
    if selected:
        lines += ["", "## Selected"]
        for item in selected:
            lines.append(f"- {item['time']} | {item['home']} vs {item['away']} | {item['market_label']} | prob {item['probability']:.2f}% | odd {item['odd']}")
    path.write_text("\n".join(lines), encoding="utf-8")


def default_target_date() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d")


def tomorrow_date() -> str:
    return (datetime.utcnow() + timedelta(days=1)).strftime("%Y-%m-%d")


def main() -> int:
    parser = argparse.ArgumentParser(description="Dry run filtri Match Center con rebuild mercati dalla cache")
    parser.add_argument("--repo-root", default=".")
    parser.add_argument("--cache-dir", default="assets/data/match-predictor")
    parser.add_argument("--cache-file", default="")
    parser.add_argument("--date", default=default_target_date())
    parser.add_argument("--max-picks", type=int, default=40)
    parser.add_argument("--output-dir", default="_match_center_filter_dry_run_singlefile")
    args = parser.parse_args()

    repo_root = Path(args.repo_root).resolve()
    target_date = str(args.date).strip()
    cache_file = resolve_cache_file(repo_root, args.cache_dir, target_date, args.cache_file.strip() or None)
    data = json.loads(cache_file.read_text(encoding="utf-8"))

    normalized = [normalize_match(item) for item in iter_match_candidates(data)]
    all_candidates: List[dict] = []
    debug_samples = []
    for match in normalized:
        candidates, reasons = apply_rules(match)
        all_candidates.extend(candidates)
        if len(debug_samples) < 25:
            debug_samples.append({
                "fixture_id": match["fixture_id"],
                "home": match["home"],
                "away": match["away"],
                "markets_found": sorted(match["markets"].keys()),
                "reasons": reasons,
                "market_preview": match["markets"],
                "flat_keys_sample": match["flat_keys_sample"],
            })

    unique_selected = choose_unique_highest_odd(all_candidates)
    selected = unique_selected[: max(0, args.max_picks)]

    out_dir = (repo_root / args.output_dir / target_date).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    base = out_dir / f"match_center_dry_run_{target_date}"

    counts_raw = Counter(item["market_label"] for item in all_candidates)
    counts_selected = Counter(item["market_label"] for item in selected)
    summary = {
        "date": target_date,
        "cache_file": str(cache_file),
        "matches_normalized": len(normalized),
        "all_candidates": len(all_candidates),
        "selected_unique_before_cap": len(unique_selected),
        "final_selected": len(selected),
        "max_picks": args.max_picks,
        "breakdown_raw": {rule["label"]: counts_raw.get(rule["label"], 0) for rule in MARKET_RULES},
        "breakdown_final": {rule["label"]: counts_selected.get(rule["label"], 0) for rule in MARKET_RULES},
    }

    by_market = {rule["id"]: [r for r in all_candidates if r.get("market_id") == rule["id"]] for rule in MARKET_RULES}
    workbook_path = base.with_name(base.name + "_full.xlsx")
    write_xlsx(workbook_path, selected, all_candidates, summary, by_market, debug_samples)

    print("=" * 72)
    print("MATCH CENTER DRY RUN SINGLE FILE")
    print(f"Date: {target_date}")
    print(f"Cache: {cache_file}")
    print("=" * 72)
    print(f"# Matches normalized: {len(normalized)}")
    print(f"# All candidates passing filters: {len(all_candidates)}")
    print(f"# Selected unique before cap: {len(unique_selected)}")
    print(f"# Final selected (cap {args.max_picks}): {len(selected)}")
    print("\n# Breakdown after cap:")
    for rule in MARKET_RULES:
        print(f"- {rule['label']}: {counts_selected.get(rule['label'], 0)}")
    print(f"\n# Output dir: {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
